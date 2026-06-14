from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.shortcuts import render
from .models import Empleado, Marcacion
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime
from django.utils import timezone
from datetime import timedelta
from django.core.management import call_command
from django.http import HttpResponse

def inicio(request):

    return render(
        request,
        "asistencia/inicio.html"
    )

def estado_empleados(request):

    empleados = []

    dentro = 0
    fuera = 0

    for empleado in Empleado.objects.all():

        ultima_marcacion = empleado.marcacion_set.order_by(
            '-fecha_hora'
        ).first()

        if ultima_marcacion:

            if ultima_marcacion.tipo in [
                'ENTRADA',
                'ENTRADA_DESCANSO',
                'ENTRADA_TE'
            ]:

                estado = "DENTRO"
                dentro += 1

            else:

                estado = "FUERA"
                fuera += 1

        else:

            estado = "FUERA"
            fuera += 1

        empleados.append({
            "id": empleado.id,
            "nombre": empleado.nombre,
            "estado": estado
        })

    return render(
        request,
        "asistencia/estado.html",
        {
            "empleados": empleados,
            "dentro": dentro,
            "fuera": fuera
        }
    )


def detalle_empleado(request, empleado_id):

    empleado = Empleado.objects.get(id=empleado_id)

    mes = request.GET.get('mes', 'todos')
    anio = request.GET.get('anio', 'todos')

    marcaciones = empleado.marcacion_set.all().order_by('-fecha_hora')

    if mes != "todos":
        marcaciones = marcaciones.filter(
            fecha_hora__month=int(mes)
        )

    if anio != "todos":
        marcaciones = marcaciones.filter(
            fecha_hora__year=int(anio)
        )

    filas = []

    for marcacion in marcaciones:

        filas.append({
            "fecha": marcacion.fecha_hora,
            "tipo": marcacion.tipo
        })

    return render(
        request,
        "asistencia/empleado.html",
        {
            "empleado": empleado,
            "filas": filas,
            "mes_actual": mes,
            "anio_actual": anio
        }
    )

def reporte_general(request):

    mes = request.GET.get('mes', 'todos')
    anio = request.GET.get('anio', 'todos')

    marcaciones = Marcacion.objects.all().order_by('-fecha_hora')

    if mes != "todos":
        marcaciones = marcaciones.filter(
            fecha_hora__month=int(mes)
        )

    if anio != "todos":
        marcaciones = marcaciones.filter(
            fecha_hora__year=int(anio)
        )

    filas = []

    for marcacion in marcaciones:

        filas.append({
            "empleado": marcacion.empleado.nombre,
            "fecha": marcacion.fecha_hora,
            "tipo": marcacion.tipo
        })

    return render(
        request,
        "asistencia/reporte.html",
        {
            "filas": filas,
            "mes_actual": mes,
            "anio_actual": anio
        }
    )

def exportar_excel(request, empleado_id):

    empleado = Empleado.objects.get(id=empleado_id)

    mes = request.GET.get('mes', 'todos')
    anio = request.GET.get('anio', 'todos')

    marcaciones = empleado.marcacion_set.all().order_by('-fecha_hora')

    if mes != "todos":
        marcaciones = marcaciones.filter(
            fecha_hora__month=int(mes)
        )

    if anio != "todos":
        marcaciones = marcaciones.filter(
            fecha_hora__year=int(anio)
        )

    wb = Workbook()
    ws = wb.active

    ws.title = "Asistencia"

    encabezados = [
        "Empleado",
        "Fecha y Hora",
        "Tipo"
    ]

    ws.append(encabezados)

    encabezado_fill = PatternFill(
        start_color="1E293B",
        end_color="1E293B",
        fill_type="solid"
    )

    encabezado_font = Font(
        color="FFFFFF",
        bold=True
    )

    for celda in ws[1]:
        celda.fill = encabezado_fill
        celda.font = encabezado_font
        celda.alignment = Alignment(horizontal="center")

    for marcacion in marcaciones:

        tipo = marcacion.tipo

        ws.append([
            empleado.nombre,
            marcacion.fecha_hora.strftime("%d/%m/%Y %H:%M:%S"),
            tipo
        ])

        fila_actual = ws.max_row

        colores = {
            "ENTRADA": "22C55E",
            "SALIDA": "EF4444",
            "ENTRADA_DESCANSO": "3B82F6",
            "SALIDA_DESCANSO": "F59E0B",
            "ENTRADA_TE": "8B5CF6",
            "SALIDA_TE": "6B7280",
        }

        color = colores.get(tipo, "FFFFFF")

        ws[f"C{fila_actual}"].fill = PatternFill(
            start_color=color,
            end_color=color,
            fill_type="solid"
        )

    for columna in ws.columns:

        longitud_maxima = 0
        letra_columna = get_column_letter(columna[0].column)

        for celda in columna:

            try:
                if len(str(celda.value)) > longitud_maxima:
                    longitud_maxima = len(str(celda.value))
            except:
                pass

        ws.column_dimensions[letra_columna].width = longitud_maxima + 5

    ws.freeze_panes = "A2"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    if mes != "todos" and anio != "todos":
        nombre_archivo = (
            f"{empleado.nombre}_{anio}_{mes}.xlsx"
        )
    else:
        nombre_archivo = (
            f"{empleado.nombre}_completo.xlsx"
        )

    response[
        'Content-Disposition'
    ] = f'attachment; filename="{nombre_archivo}"'

    wb.save(response)

    return response

def exportar_reporte_excel(request):

    mes = request.GET.get('mes', 'todos')
    anio = request.GET.get('anio', 'todos')

    marcaciones = Marcacion.objects.all().order_by('-fecha_hora')

    if mes != "todos":
        marcaciones = marcaciones.filter(
            fecha_hora__month=int(mes)
        )

    if anio != "todos":
        marcaciones = marcaciones.filter(
            fecha_hora__year=int(anio)
        )

    wb = Workbook()
    ws = wb.active

    ws.title = "Reporte General"

    encabezados = [
        "Empleado",
        "Fecha y Hora",
        "Tipo"
    ]

    ws.append(encabezados)

    encabezado_fill = PatternFill(
        start_color="1E293B",
        end_color="1E293B",
        fill_type="solid"
    )

    encabezado_font = Font(
        color="FFFFFF",
        bold=True
    )

    for celda in ws[1]:
        celda.fill = encabezado_fill
        celda.font = encabezado_font
        celda.alignment = Alignment(horizontal="center")

    for marcacion in marcaciones:

        tipo = marcacion.tipo

        ws.append([
            marcacion.empleado.nombre,
            marcacion.fecha_hora.strftime("%d/%m/%Y %H:%M:%S"),
            tipo
        ])

        fila_actual = ws.max_row

        colores = {
            "ENTRADA": "22C55E",
            "SALIDA": "EF4444",
            "ENTRADA_DESCANSO": "3B82F6",
            "SALIDA_DESCANSO": "F59E0B",
            "ENTRADA_TE": "8B5CF6",
            "SALIDA_TE": "6B7280",
        }

        color = colores.get(tipo, "FFFFFF")

        ws[f"C{fila_actual}"].fill = PatternFill(
            start_color=color,
            end_color=color,
            fill_type="solid"
        )

    for columna in ws.columns:

        longitud_maxima = 0
        letra_columna = get_column_letter(columna[0].column)

        for celda in columna:

            try:
                if len(str(celda.value)) > longitud_maxima:
                    longitud_maxima = len(str(celda.value))
            except:
                pass

        ws.column_dimensions[letra_columna].width = longitud_maxima + 5

    ws.freeze_panes = "A2"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    if mes != "todos" and anio != "todos":
        nombre_archivo = f"Reporte_{anio}_{mes}.xlsx"
    else:
        nombre_archivo = "Reporte_Completo.xlsx"

    response[
        'Content-Disposition'
    ] = f'attachment; filename="{nombre_archivo}"'

    wb.save(response)

    return response

@csrf_exempt
def recibir_datos_adms(request):

    print("================================")
    print("NUEVA PETICION ADMS")
    print("METODO:", request.method)
    print("URL:", request.get_full_path())
    print("HEADERS:", dict(request.headers))
    print("================================")

    numero_serie = request.GET.get(
        'SN',
        'Desconocido'
    )

    tabla = request.GET.get(
        'table',
        'DESCONOCIDA'
    )

    ip_remota = request.META.get(
        'REMOTE_ADDR',
        'DESCONOCIDA'
    )

    fecha_recepcion = datetime.now().strftime(
        "%d/%m/%Y %H:%M:%S"
    )

    if request.method == 'GET':

        print(
            f"📡 [ADMS] El reloj {numero_serie} se acaba de conectar."
        )

        return HttpResponse("OK\n")

    elif request.method == 'POST':

        datos_crudos = request.body.decode(
            'utf-8',
            errors='ignore'
        )


        print(
            f"\n========================================"
        )

        print(
            f"✅ ¡NUEVO REGISTRO DESDE EL RELOJ {numero_serie}!"
        )

        print(
            datos_crudos
        )

        print(
            "========================================\n"
        )

        # SOLO procesamos ATTLOG
        if tabla != "ATTLOG":
            return HttpResponse("OK\n")

        try:

            partes = datos_crudos.strip().split()

            if len(partes) < 4:
                return HttpResponse("OK\n")

            id_biometrico = int(partes[0])

            fecha_hora = datetime.strptime(
                f"{partes[1]} {partes[2]}",
                "%Y-%m-%d %H:%M:%S"
            )
            if numero_serie == "VGU6243400121":
                fecha_hora = fecha_hora - timedelta(hours=1)

            codigo_estado = partes[3]

            mapa_estados = {
                "0": "ENTRADA",
                "1": "SALIDA",
                "2": "SALIDA_DESCANSO",
                "3": "ENTRADA_DESCANSO",
                "4": "ENTRADA_TE",
                "5": "SALIDA_TE",
            }

            tipo = mapa_estados.get(
                codigo_estado,
                "ENTRADA"
            )

            empleado = Empleado.objects.get(
                id_biometrico=id_biometrico
            )

            Marcacion.objects.create(
                empleado=empleado,
                fecha_hora=fecha_hora,
                tipo=tipo
            )

            print(
                f"✅ Marcación guardada para {empleado.nombre}"
            )

        except Exception as e:

            print(
                f"❌ Error procesando marcación: {e}"
            )

        return HttpResponse("OK\n")

    return HttpResponse("OK\n")



def importar_datos(request):

    try:
        call_command(
            'loaddata',
            'respaldo_completo.json'
        )

        return HttpResponse(
            "Datos importados correctamente"
        )

    except Exception as e:

        return HttpResponse(
            f"Error: {e}"
        )
